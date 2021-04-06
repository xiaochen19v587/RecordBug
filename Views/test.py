from openpyxl import load_workbook

def openxl():
    wb = load_workbook('/home/user/测试用例/APA测试用例v14 (第9个复件).xlsx')
    wb.active
    worksheet = wb["车位检测测试"]
    worksheet.cell(2,9).value = 'test'
    wb.save('/home/user/测试用例/APA测试用例v14 (第9个复件).xlsx')

if __name__ == '__main__':
    openxl()