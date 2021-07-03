#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import re
import subprocess
import xlrd
import time
from openpyxl import load_workbook
from time import sleep
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from Ui_RecordInfo import Ui_RecordBug
from Ui_PullFile import Ui_PullFile
from Ui_ChoosePull import Ui_ChoosePull
from Ui_AdbPull import Ui_AdbPull
from Ui_ScpPull import Ui_ScpPull
from Ui_ChoosePush import Ui_ChoosePush
from Ui_AdbPush import Ui_AdbPush
from Ui_ScpPush import Ui_ScpPush
from Ui_OpenRviz import Ui_OpenRviz
from Ui_BrushSoc import Ui_BrushSoc


class Record_Info_Views(QMainWindow, Ui_RecordBug):
    '''
        主界面
    '''

    def __init__(self):
        '''
        初始化主界面,设置定时器用于显示时间,设置定时器状态time_count,设置第二界面问题保存标志pushButton_savecount
        time_count 1 开始 0 暂停
        pushButton_savecount 0 初始状态 1 2 3 对应当前测试次数 4 添加了问题
        '''
        super().__init__()
        self.create_log_daily = CREATE_LOG_DAILY()
        self.setupUi(self)
        self.timer = QTimer(self)
        self.pushButton_savecount = 0
        self.fileName = ''
        self.old_fileName_choose = ''
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())
        self.initUI()

    def initUI(self):
        '''
        设置问题保存路径,给按钮绑定事件,给下拉框绑定事件,给列表和表格绑定点击事件
        '''
        self.test_type()
        self.file_path = '/home/user/Data/Record_Info'
        self.pushButton.clicked.connect(self.get_info)
        self.pushButton_2.clicked.connect(self.clear_all)
        self.pushButton_3.clicked.connect(self.pull_file)
        self.pushButton_4.clicked.connect(self.choose_push)
        self.pushButton_5.clicked.connect(self.save_update)
        self.pushButton_6.clicked.connect(self.save_info)
        self.pushButton_7.clicked.connect(self.choose_xlsx_file)
        self.pushButton_8.clicked.connect(self.show_time)
        self.pushButton_9.clicked.connect(self.choose_pull)
        self.pushButton_10.clicked.connect(self.test_save_fail)
        self.pushButton_11.clicked.connect(self.first_test)
        self.pushButton_12.clicked.connect(self.second_test)
        self.pushButton_13.clicked.connect(self.third_test)
        self.pushButton_14.clicked.connect(self.show_time)
        self.pushButton_15.clicked.connect(self.test_save_pass)
        self.pushButton_16.clicked.connect(self.save_to_excel)
        self.pushButton_17.clicked.connect(self.open_rviz)
        self.pushButton_18.clicked.connect(self.test_save_info)
        self.pushButton_19.clicked.connect(self.brush_soc)
        self.pushButton_20.clicked.connect(self.get_carinfo)
        self.pushButton_21.clicked.connect(self.get_test_progress)
        self.comboBox.currentTextChanged.connect(self.change_sheet)
        self.comboBox_2.currentIndexChanged.connect(self.test_type)
        self.comboBox_5.currentTextChanged.connect(self.change_items)
        self.tableWidget.clicked.connect(self.table_click)
        self.listWidget.itemClicked.connect(self.update_info)

# 第一界面
    def get_info(self):
        '''
        car_name 车辆信息,question_time 问题时间,test_type 测试类型,test_path 测试路线,question_place 问题位置,
        fault_rate 故障频率 question_describe 问题描述,err_info 详细问题记录
        '''
        self.create_log_daily.function_start_log("get_info")
        car_name = self.label_17.text()
        question_time = self.label_3.text()
        test_type = self.comboBox_2.currentText()
        test_path = self.lineEdit.text()
        question_place = self.lineEdit_2.text()
        fault_rate = self.comboBox_3.currentText()
        question_describe = self.comboBox_4.currentText()+' ' + \
            self.plainTextEdit.toPlainText()
        err_info = "车辆信息 : {}\n问题时间 : {}\n测试类型 : {}\n测试路线 : {}\n问题位置 : {}\n故障频率 : {}\n问题描述 : {}\n".format(
            car_name, question_time, test_type, test_path, question_place, fault_rate, question_describe)
        get_info_date = "车辆信息 : {}问题时间 : {}测试类型 : {}测试路线 : {}问题位置 : {}故障频率 : {}问题描述 : {}".format(
            car_name, question_time, test_type, test_path, question_place, fault_rate, question_describe)
        self.create_log_daily.function_info_log(
            "get_info", get_info_date)
        self.record_info(err_info)
        self.create_log_daily.function_close_log("get_info")

    def get_carinfo(self):
        '''
        创建default.xml文件文件保存路径,通过adb方式拉取default.xml,读取文件中的车辆信息
        '''
        self.create_log_daily.function_start_log("get_car_info")
        self.label_10.setText('')
        res = Mkdir_Path_Views().mkdir_dir_path('/home/user/Data/car_instance/')
        if res:
            self.create_log_daily.function_info_log(
                "get_carinfo", "car_instance Folder created successfully")
            res = subprocess.call(
                'timeout 1 adb pull /data/zros/res/car_instance/default.xml /home/user/Data/car_instance/', shell=True)
            if res:
                self.create_pop('adb is not connect')
                self.create_log_daily.function_info_log(
                    "get_info", "adb is not connect")
            else:
                data = os.popen('cat /home/user/Data/car_instance/default.xml')
                car_info = re.findall(
                    '<car_instance>(.*?)</car_instance>', data.read())[0]
                self.create_log_daily.function_info_log(
                    "get_info", "car_instance File read succeeded")
                self.label_17.setText(car_info)
                self.create_log_daily.function_info_log(
                    "get_info", "default.xml is set succeeded")
        else:
            self.create_log_daily.function_info_log(
                "get_carinfo", "car_instance Failed to create folder")
            self.create_pop('创建/home/user/Data/car_instance/文件夹失败')
        self.create_log_daily.function_close_log("get_info")

    def record_info(self, err_info):
        '''
        err_info:问题信息
        记录问题,重置右侧内容,设置记录成功的标记
        '''
        self.create_log_daily.function_start_log("record_info")
        self.listWidget.addItem(err_info)
        self.clear_all()
        self.create_log_daily.function_info_log(
            "record_info", "Reset complete")
        self.label_10.setText("记录成功")
        self.create_log_daily.function_info_log(
            "record_info", "Record succeeded")
        self.create_log_daily.function_close_log("record_info")

    def clear_all(self):
        '''
        重置界面右侧填写的所有信息
        '''
        self.create_log_daily.function_start_log("clear_all")
        self.comboBox_2.setCurrentIndex(0)
        self.comboBox_3.setCurrentIndex(0)
        self.plainTextEdit.setPlainText('')
        self.lineEdit.setText('')
        self.lineEdit_2.setText('')
        self.label_3.setText("")
        self.label_10.setText("")
        self.label_13.setText("")
        self.create_log_daily.function_info_log(
            "clear_all", "Reset complete")
        self.create_log_daily.function_close_log("clear_all")

    def show_time(self):
        '''
        显示时间
        '''
        self.create_log_daily.function_start_log("show_time")
        datetime = QDateTime.currentDateTime()
        years = datetime.toString().split(' ')[-1]
        months = datetime.toString().split(' ')[1][0:2]
        days = datetime.toString().split(' ')[2]
        times = datetime.toString().split(' ')[3]
        self.label_3.setText(
            '{}年{}{}日 {}'.format(years, months, days, times))
        self.label_13.setText(
            '{}年{}{}日 {}'.format(years, months, days, times))
        self.create_log_daily.function_info_log(
            "show_time", "set time is {}年{}{}日 {}".format(years, months, days, times))
        self.create_log_daily.function_close_log("show_time")

    def keyPressEvent(self, event):
        '''
        键盘事件监听函数, Esc 暂停和开始时间,Ctrl+c 关闭主界面窗口
        '''
        self.create_log_daily.function_start_log("keyPressEvent")
        if event.key() == Qt.Key_Escape:
            self.create_log_daily.function_info_log(
                "keyPressEvent", "Esc is pressed")
            self.show_time()
        if event.modifiers() == Qt.ControlModifier and event.key() == Qt.Key_C:
            self.create_log_daily.function_info_log(
                "keyPressEvent", "Ctrl and C is pressed")
            self.closeEvent(event)
        self.create_log_daily.function_close_log("keyPressEvent")

    def save_info(self):
        '''
        将左侧问题记录中的所有信息保存到文件中,清除问题记录列表中所有信息
        '''
        self.create_log_daily.function_start_log("save_info")
        if self.listWidget.count():
            res = Mkdir_Path_Views().mkdir_dir_path(self.file_path)
            if res:
                self.create_log_daily.function_info_log(
                    "save_info", "Record_Info Folder created successfully")
                with open('{}/bug_file.txt'.format(self.file_path), 'a') as f:
                    for count in range(self.listWidget.count()):
                        f.write(self.listWidget.item(count).text()+'\n')
                self.label_10.setText(
                    "问题记录已保存到{}/bug_file.txt".format(self.file_path))
                self.create_log_daily.function_info_log(
                    "save_info", "bug_file.txt was written successfully")
                self.listWidget.clear()
            else:
                self.create_log_daily.function_info_log(
                    "save_info", "Record_Info Failed to create folder")
                self.create_pop("创建{}文件夹失败".format(self.file_path))
        else:
            self.create_log_daily.function_info_log(
                "save_info", "is no problem record")
            self.create_pop("没有问题记录")
        self.create_log_daily.function_close_log("save_info")

    def test_type(self):
        '''
        设置测试类型
        '''
        self.create_log_daily.function_start_log("test_type")
        self.comboBox_4.clear()
        if self.comboBox_2.currentText() == "接驾":
            err_list = ['接驾未完成', '过弯靠外侧', '过弯靠内侧', '障碍物误检测', '定位飘了', '其他']
        elif self.comboBox_2.currentText() == "泊车":
            err_list = ['泊车歪了', '挂p档', '泊车未完成', '障碍物误检测', '定位飘了', '其他']
        elif self.comboBox_2.currentText() == "跨层":
            err_list = ['未进线控', '路径规划时间过长', '溜坡', '障碍物误检测', '定位飘了', '其他']
        elif self.comboBox_2.currentText() == "避障":
            err_list = ['障碍物未识别', '障碍物误检测', '障碍物检测成功,未刹车',
                        '障碍物检测成功并避障,障碍物移除之后不能继续接驾', '其他']
        elif self.comboBox_2.currentText() == "定位":
            err_list = ['定位失败', '定位成功地图未切换', '其他']
        self.create_log_daily.function_info_log("test_type", "Current test type is {},err_list is {}".format(
            self.comboBox_2.currentText(), err_list))
        self.comboBox_4.addItems(err_list)
        self.create_log_daily.function_close_log("test_type")

    def update_info(self, item):
        '''
        将左侧问题记录中的信息显示在右侧对应位置
        '''
        self.create_log_daily.function_start_log('update_info')
        err_describe = ''
        self.comboBox_2.setCurrentText(re.findall(
            '测试类型 : (.*?)\n', str(item.text()))[0])
        self.lineEdit.setText(re.findall(
            '测试路线 : (.*?)\n', str(item.text()))[0])
        self.lineEdit_2.setText(re.findall(
            '问题位置 : (.*?)\n', str(item.text()))[0])
        self.comboBox_3.setCurrentText(re.findall(
            '故障频率 : (.*?)\n', str(item.text()))[0])
        self.comboBox_4.setCurrentText(re.findall(
            '问题描述 : (.*?)\n', str(item.text()))[0].split(' ')[0])
        for i in range(1, len(re.findall('问题描述 : (.*?)\n', str(item.text()))[0].split(' '))):
            err_describe += re.findall('问题描述 : (.*?)\n',
                                       str(item.text()))[0].split(' ')[i] + ' '
        self.plainTextEdit.setPlainText(err_describe)
        self.create_log_daily.function_info_log(
            "update_info", "Modify problem record")
        self.create_log_daily.function_close_log("update_info")

    def save_update(self):
        '''
        保存修改后的信息,将右侧各个问题描述记录保存在右侧问题记录列表中
        '''
        self.create_log_daily.function_start_log("save_update")
        if self.listWidget.count():
            self.create_log_daily.function_info_log(
                "save_update", "listWidget is empty")
            if self.listWidget.currentItem():
                self.create_log_daily.function_info_log(
                    "save_update", "currentItem has been selected")
                car_name = self.label_17.text()
                question_time = re.findall(
                    '问题时间 : (.*?)\n', self.listWidget.currentItem().text())[0]
                test_type = self.comboBox_2.currentText()
                test_path = self.lineEdit.text()
                question_place = self.lineEdit_2.text()
                fault_rate = self.comboBox_3.currentText()
                question_describe = self.comboBox_4.currentText()+' ' + \
                    self.plainTextEdit.toPlainText()
                item = "车辆信息 : {}\n问题时间 : {}\n测试类型 : {}\n测试路线 : {}\n问题位置 : {}\n故障频率 : {}\n问题描述 : {}\n".format(
                    car_name, question_time, test_type, test_path, question_place, fault_rate, question_describe)
                self.listWidget.insertItem(self.listWidget.currentRow(), item)
                self.listWidget.takeItem(self.listWidget.currentRow())
                self.clear_all()
            else:
                self.create_log_daily.function_info_log(
                    "save_update", "not selected currentItem")
                self.create_pop('请选择要修改的记录')
        else:
            self.create_log_daily.function_info_log(
                "save_update", "listWidget is not empty")
            self.create_pop('没有问题记录')
        self.create_log_daily.function_close_log("save_update")

    def create_pop(self, info):
        '''
        设置提醒弹窗
        '''
        self.create_log_daily.function_start_log("create_pop")
        msg_box = QMessageBox(QMessageBox.Warning, '提示', info)
        msg_box.exec_()
        self.create_log_daily.function_close_log("create_pop")

# 第一界面

# 第二界面
    def choose_xlsx_file(self):
        '''
        生成文件选择窗口
        根据sheet名为测试用例读取表格
        '''
        self.create_log_daily.function_start_log("choose_xlsx_file")
        self.fileName_choose, filetype = QFileDialog.getOpenFileName(self,
                                                                     "选取文件",
                                                                     "/home/user/",
                                                                     "Text Files (*.xls);Text Files (*.xlsx)")
        if self.fileName_choose == "":
            self.create_log_daily.function_info_log(
                "choose_xlsx_file", "choose xlsx file is None")
            if self.old_fileName_choose:
                self.fileName_choose = self.old_fileName_choose
            return
        else:
            self.old_fileName_choose = self.fileName_choose
        self.create_log_daily.function_info_log(
            "choose_xlsx_file", "choose xlsx file is {}".format(self.fileName_choose))

        # 创建新的excel对象
        table_data = xlrd.open_workbook(self.fileName_choose)
        # 设置comboBox条目
        self.comboBox.clear()
        self.comboBox.addItems(table_data.sheet_names())
        self.create_log_daily.function_close_log("choose_xlsx_file")

    def change_sheet(self):
        self.create_log_daily.function_start_log("change_sheet")
        # 测试用例选中的sheet发生变化
        sheet_name = self.comboBox.currentText()
        self.create_log_daily.function_info_log(
            "change_sheet", "choose sheet is {}".format(sheet_name))
        if sheet_name == '':
            return
        self.label_19.setText('')
        self.change_table(sheet_name)
        self.create_log_daily.function_close_log("change_sheet")

    def change_items(self):
        # 测试项发生变化
        self.create_log_daily.function_start_log("change_items")
        self.create_log_daily.function_info_log(
            "change_items", "current item is {}".format(self.comboBox_5.currentText()))
        if self.comboBox_5.currentText() != "测试项":
            self.show_table(1, 1)
        else:
            self.show_table(1, 0)
        self.create_log_daily.function_close_log("change_items")

    def change_table(self, sheet_name):
        '''
            表格需要发生变化; 
            sheet_name:需要打开的sheet
        '''
        self.create_log_daily.function_start_log("change_table")
        table_data = xlrd.open_workbook(self.fileName_choose)
        self.create_log_daily.function_info_log(
            "change_table", "current choose xlsx file is {}".format(self.fileName_choose))
        try:
            self.table = table_data.sheet_by_name(sheet_name)
        except:
            self.open_xlsx(1)
        else:
            self.open_xlsx(0)
            self.set_tablewidget_header()
            self.show_table(1, 0)
        self.create_log_daily.function_close_log("change_table")

    def open_xlsx(self, method_code):
        '''
        读取xlsx文件中信息
        生成id case step result列表
        '''
        self.create_log_daily.function_start_log("open_xlsx")
        if method_code:
            return
        self.id_list = []
        self.case_list = []
        self.step_list = []
        self.result_list = []
        self.all_itmes_list = []
        self.items_list = []
        # 生成id_list,all_itmes_list,case_list,step_list,result_list
        self.create_log_daily.function_info_log(
            "open_xlsx", "generate id_list,all_itmes_list,case_list,step_list,result_list")
        for cols_index in range(self.table.ncols):
            cols_name = self.table.cell_value(0, cols_index)
            if cols_name == "测试用例ID":
                self.create_log_daily.function_info_log(
                    "list_append_value", "Building generate id_list")
                self.list_append_value(self.id_list, cols_index)
            elif cols_name == '测试项':
                self.create_log_daily.function_info_log(
                    "list_append_value", "Building generate all_itmes_list")
                self.list_append_value(self.all_itmes_list, cols_index)
            elif cols_name == '测试用例':
                self.create_log_daily.function_info_log(
                    "list_append_value", "Building case_list id_list")
                self.list_append_value(self.case_list, cols_index)
            elif cols_name == '测试步骤':
                self.create_log_daily.function_info_log(
                    "list_append_value", "Building generate step_list")
                self.list_append_value(self.step_list, cols_index)
            elif cols_name == '期望结果':
                self.create_log_daily.function_info_log(
                    "list_append_value", "Building generate result_list")
                self.list_append_value(self.result_list, cols_index)
        # 生成所有测试项的非空列表
        self.items_list.append('测试项')
        for itme_index in range(1, len(self.all_itmes_list)):
            if self.all_itmes_list[itme_index]:
                self.items_list.append(self.all_itmes_list[itme_index])
        self.create_log_daily.function_close_log("open_xlsx")

    def list_append_value(self, cols_name_list, cols_index):
        '''
        将表格中对应信息添加进对应列表
        '''
        self.create_log_daily.function_start_log("list_append_value")
        for rows_index in range(self.table.nrows):
            cols_name_list.append(
                self.table.cell_value(rows_index, cols_index))
        self.create_log_daily.function_close_log("list_append_value")

    def set_tablewidget_header(self):
        '''
            设置tablewidget头部信息
        '''
        self.create_log_daily.function_start_log("set_tablewidget_header")
        self.comboBox_5.clear()
        self.comboBox_5.addItems(self.items_list)
        # 如果id_list,case_list,step_list,result_list都为空,则清空所有内容(选择测试用例格式不正确)
        if not self.id_list and not self.case_list and not self.step_list and not self.result_list:
            self.create_log_daily.function_info_log(
                "set_tablewidget_header", "current choose xlsx file format is err")
            if self.fileName == self.old_fileName_choose:
                self.create_log_daily.function_info_log(
                    "set_tablewidget_header", "current choose xlsx file is not changed,choose sheet is changed")
                # 选择的excel文件没有变化,sheet发生变化
                self.create_pop('选择的sheet格式错误')
                self.comboBox.setCurrentIndex(0)
                self.comboBox_5.setCurrentIndex(0)
            else:
                self.create_log_daily.function_info_log(
                    "set_tablewidget_header", "current choose xlsx file is changed")
                # 选择的excel文件发生变化
                self.create_pop('选择的excel文件格式错误')
                self.tableWidget.setRowCount(0)
                self.tableWidget.setColumnCount(0)
                self.comboBox.clear()
                self.comboBox_5.clear()
            return
        else:
            self.create_log_daily.function_info_log(
                "set_tablewidget_header", "current choose xlsx file format is right")
            self.create_log_daily.function_info_log(
                "set_tablewidget_header", "current choose xlsx file is {}".format(self.old_fileName_choose))
            self.fileName = self.old_fileName_choose
            # 设置tablewidget的列数
            self.tableWidget.setColumnCount(3)
            # 设置tablewidget顶部标题
            self.tableWidget.setHorizontalHeaderLabels(
                ['测试用例ID', '测试用例', '测试次数'])
            # 设置行号隐藏
            self.tableWidget.verticalHeader().setVisible(False)
            # 设置整行选中
            self.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.create_log_daily.function_close_log("set_tablewidget_header")

    def show_table(self, sheet_change_code, item_change_code):
        '''
            sheet_change_code: 1 sheet发生改变 0 sheet没有发生变化;
            item_change_code: 1 测试项发生改变 0 测试项没有发生变化
        '''
        self.create_log_daily.function_start_log("show_table")
        if sheet_change_code:
            self.create_log_daily.function_info_log(
                "show_table", "current choose sheet or item is changed")
            # sheet发生变化或者测试项发生变化
            self.textBrowser.setText('')
            self.textBrowser_2.setText('')
            self.textBrowser_3.setText('')
            self.textBrowser_6.setText('')
        self.plainTextEdit_2.setPlainText('')
        self.textBrowser_4.setPlainText('')
        self.table_case_id = {}
        self.pushButton_savecount = 0
        self.pushButton_count = 0
        self.create_log_daily.function_info_log(
            "show_table", "set table state")
        # 设置表格具体的内容和行数
        self.show_table_content(sheet_change_code, item_change_code)
        # 设置表格内容不可修改(会导致表格内容显示不完全)
        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # 设置表格大小根据内容自适应
        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.resizeRowsToContents()
        # 设置表格铺满
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.horizontalHeader().setCascadingSectionResizes(
            QHeaderView.ResizeToContents)
        self.create_log_daily.function_close_log("show_table")

    def show_table_content(self, sheet_change_code, item_change_code):
        '''
        sheet_change_code: 1 sheet发生改变 0 sheet没有发生变化
        item_change_code: 1 测试项发生改变 0 测试项没有发生变化
        '''
        self.create_log_daily.function_start_log("show_table_content")
        if item_change_code:
            self.create_log_daily.function_info_log(
                "show_table_content", "current choose item is changed")
            # 测试项发生变化
            self.tableWidget.setRowCount(0)
            self.get_item_count()
            table_length = self.start_stop_index[1]-self.start_stop_index[0]
            start_index = self.start_stop_index[0]
            stop_index = self.start_stop_index[1]
        else:
            self.create_log_daily.function_info_log(
                "show_table_content", "current choose item is not changed")
            # 测试项没有发生改变
            table_length = len(self.case_list)-1
            start_index = 1
            stop_index = len(self.case_list)
        count = 0
        # 设置行数
        self.tableWidget.setRowCount(table_length)
        self.create_log_daily.function_info_log(
            "show_table_content", "current set row count is {}".format(table_length))
        # 设置表格内容
        for case_index in range(start_index, stop_index):
            self.table_case_id[self.id_list[case_index]] = 0
            # 测试用例ID
            self.tableWidget.setItem(
                count, 0, QTableWidgetItem(self.id_list[case_index]))
            # 测试用例
            self.tableWidget.setItem(
                count, 1, QTableWidgetItem(self.case_list[case_index].split('\n')[0]))
            # 如果sheet发生变化
            # 清空tablewidget测试次数列
            if sheet_change_code:
                self.tableWidget.setItem(count, 2, QTableWidgetItem(''))
            count += 1
        self.create_log_daily.function_close_log("show_table_content")

    def get_item_count(self):
        '''
            获取选中的测试项的数量,生成self.start_stop_index元组,
            第一个元素代表当前选中的测试项的索引,第二个元素代表下一个测试项的索引
        '''
        self.create_log_daily.function_start_log("get_item_count")
        start_index = 0
        stop_index = 0
        next_item_name = 0
        # 获取start_index
        self.create_log_daily.function_info_log(
            "get_item_count", "Getting start_index")
        for i in range(1, len(self.all_itmes_list)):
            if self.comboBox_5.currentText() == self.all_itmes_list[i]:
                start_index = i
        # 获取next_item_name
        self.create_log_daily.function_info_log(
            "get_item_count", "Getting next_item_name")
        for i in range(1, len(self.items_list)):
            if self.comboBox_5.currentText() == self.items_list[i]:
                if i+1 == len(self.items_list):
                    stop_index = len(self.all_itmes_list)
                else:
                    next_item_name = self.items_list[i+1]
                    # 获取stop_index
                    for i in range(1, len(self.all_itmes_list)):
                        if next_item_name == self.all_itmes_list[i]:
                            stop_index = i
        self.start_stop_index = (start_index, stop_index)
        self.create_log_daily.function_close_log("get_item_count")

    def get_test_progress(self):
        self.create_log_daily.function_start_log("get_test_progress")
        try:
            self.id_list
        except:
            return
        if self.id_list:
            self.create_log_daily.function_info_log(
                "get_test_progress", "current choose testcase list size is {}".format(len(self.id_list)))
            test_content_list = []
            # 获取新的测试用例文件对象
            table_data = xlrd.open_workbook(self.fileName_choose)
            table = table_data.sheet_by_name(self.comboBox.currentText())
            # 获取当前选择的测试用例文件中测试结果的列数
            for cols_index in range(table.ncols):
                cols_name = table.cell_value(0, cols_index)
                if cols_name == '测试结果':
                    test_res_cols = cols_index
            # 获取当前选择的测试用例文件中测试结果的内容
            for rows_index in range(table.nrows):
                rows_content = table.cell_value(rows_index, test_res_cols)
                if rows_content:
                    test_content_list.append(rows_content)
            self.label_19.setText(
                str(len(test_content_list)-1)+'/'+str(len(self.id_list)-1))
            self.create_log_daily.function_info_log(
                "get_test_progress", "current choose testcase test progress is {}".format(len(test_content_list)-1))
        self.create_log_daily.function_close_log("get_test_progress")

    def table_click(self):
        '''
        切换表格,判断pushButton_savecount,如果为0,直接切换(问题记录框没有信息);如果为1 2 3,清空问题记录框中的信息并切换(只选择了测试次数没有添加测试问题);
        如果为4,弹出提示框(问题记录框中已添加了问题)
        '''
        self.create_log_daily.function_start_log("table_click")
        if self.pushButton_savecount == 0:
            pass
        elif self.pushButton_savecount in [1, 2, 3]:
            self.textBrowser_4.setText('')
        else:
            if self.textBrowser_4.toPlainText():
                quitMsgBox = QMessageBox()
                quitMsgBox.setWindowTitle('确认提示')
                quitMsgBox.setText('你还没有保存问题,继续切换测试用例吗?')
                quitMsgBox.setStandardButtons(
                    QMessageBox.Yes | QMessageBox.No)
                buttonY = quitMsgBox.button(QMessageBox.Yes)
                buttonY.setText('保存并切换')
                buttonN = quitMsgBox.button(QMessageBox.No)
                buttonN.setText('不保存切换')
                quitMsgBox.exec_()
                self.create_log_daily.function_info_log(
                    "table_click", "current cliecked button is {}".format(quitMsgBox.clickedButton()))
                if quitMsgBox.clickedButton() == buttonY:
                    self.save_to_excel()
                    return
        self.pushButton_count = 0
        self.show_test_info()
        self.err_list()
        self.plainTextEdit_2.setPlainText('')
        self.textBrowser_4.setPlainText('')
        self.pushButton_savecount = 0
        self.create_log_daily.function_close_log("table_click")

    def show_test_info(self):
        '''
        显示测试信息,根据测试用例id读取测试步骤列表step_list和期望结果列表result_list中的内容添加到对应的显示框
        '''
        self.create_log_daily.function_start_log("show_test_info")
        case_list_data = ''
        test_id = self.tableWidget.selectedItems()[0].text()
        if not test_id:
            self.create_log_daily.function_info_log(
                "show_test_info", "current tableWidget.selectedItems test_id is None")
            # 如果没有测试用例id,将测试用例,测试步骤,期望结果,测试结果全部清空
            self.textBrowser.setText('')
            self.textBrowser_2.setText('')
            self.textBrowser_3.setText('')
            self.textBrowser_6.setText('')
        else:
            self.create_log_daily.function_info_log(
                "show_test_info", "current tableWidget.selectedItems test_id is {}".format(test_id))
            # 如果有测试用例id,在对应控件中显示对应的信息
            for i in range(1, len(self.id_list)):
                if self.id_list[i] == test_id:
                    self.textBrowser.setText("测试用例:\n" + self.case_list[i])
                    self.textBrowser_2.setText("期望结果:\n" + self.result_list[i])
                    self.textBrowser_3.setText("测试步骤:\n" + self.step_list[i])
        self.create_log_daily.function_close_log("show_test_info")

    def err_list(self):
        '''
        问题列表,读取保存文件,根据当前选择的测试用例ID将文件中记录的问题分离,显示在对应的显示框
        将table_case_id字典中索引对应的值修改为文件中的次数
        '''
        self.create_log_daily.function_start_log("err_list")
        test_id = self.tableWidget.selectedItems()[0].text()
        if not test_id:
            self.create_log_daily.function_info_log(
                "err_list", "current tableWidget.selectedItems test_id is None")
            self.create_pop("当前选择测试用例ID为空")
        else:
            self.create_log_daily.function_info_log(
                "err_list", "current tableWidget.selectedItems test_id is {}".format(test_id))
            # 获取测试用例id对应的测试结果的行号和列号
            res_rows_index, res_cols_index = self.get_rows_cols(test_id)
            # 加载新的excel对象
            table_data = xlrd.open_workbook(self.fileName_choose)
            table = table_data.sheet_by_name(self.comboBox.currentText())
            # 获取excel对象中当前选中测试用例的测试结果
            info = table.cell_value(res_rows_index, res_cols_index)
            if not info:
                self.create_log_daily.function_info_log(
                    "err_list", "current tableWidget.selectedItems test_res is None")
                # 如果数据为空
                self.tableWidget.setItem(
                    self.tableWidget.currentRow(), 2, QTableWidgetItem('还没有进行测试'))
                self.table_case_id[self.tableWidget.selectedItems()[
                    0].text()] = 0
                self.color_change(Qt.black)
                self.textBrowser_6.setText('')
                # 如果数据存在
            elif info[-3:-2] in ['1', '2', '3']:
                self.create_log_daily.function_info_log(
                    "err_list", "current tableWidget.selectedItems test_res is Exist")
                self.table_case_id[self.tableWidget.selectedItems()[
                    0].text()] = int(info[-3:-2])
                self.tableWidget.setItem(
                    self.tableWidget.currentRow(), 2, QTableWidgetItem('已进行过{}次测试'.format(self.table_case_id[self.tableWidget.selectedItems()[
                        0].text()])))
                self.color_change(Qt.red)
                self.textBrowser_6.setText(info)
            else:
                self.textBrowser_6.setText(info)
        self.create_log_daily.function_close_log("err_list")

    def test_save_info(self):
        '''
        记录测试问题,将右侧问题输入框中的内容记录在左侧问题显示框
        '''
        self.create_log_daily.function_start_log("test_save_info")
        try:
            self.tableWidget.selectedItems()[0].text()
        except:
            return
        self.create_log_daily.function_info_log(
            "test_save_info", "current pushButton_count is {}".format(self.pushButton_count))
        if self.pushButton_count:
            self.create_log_daily.function_info_log(
                "test_save_pass", "current test_info is {}".format(self.plainTextEdit_2.toPlainText()))
            if self.plainTextEdit_2.toPlainText():
                self.change_info(self.plainTextEdit_2.toPlainText())
            else:
                self.create_log_daily.function_info("test_save_info","current input test questions is None")
                self.create_pop('请输入测试问题描述')
                self.create_log_daily.function_close_log("test_save_info")
                return
        else:
            self.choose_test_pop()

    def test_save_pass(self):
        '''
        记录测试问题为pass,生成一条问题为pass的记录,记录在左侧问题显示框
        '''
        self.create_log_daily.function_start_log("test_save_pass")
        try:
            self.tableWidget.selectedItems()[0].text()
        except:
            return
        self.create_log_daily.function_info_log(
            "test_save_pass", "current pushButton_count is {}".format(self.pushButton_count))
        if self.pushButton_count:
            self.create_log_daily.function_info_log(
                "test_save_pass", "current test_info is PASS")
            self.change_info('PASS')
        else:
            self.choose_test_pop()
        self.create_log_daily.function_close_log("test_save_pass")

    def test_save_fail(self):
        '''
        记录测试问题为fail,生成一条问题为fail的记录,记录在左侧问题显示框
        '''
        self.create_log_daily.function_start_log("test_save_fail")
        try:
            self.tableWidget.selectedItems()[0].text()
        except:
            return
        self.create_log_daily.function_info_log(
            "test_save_fail", "current pushButton_count is {}".format(self.pushButton_count))
        if self.pushButton_count:
            self.create_log_daily.function_info_log(
                "test_save_fail", "current test_info is FAIL")
            self.change_info('FAIL')
        else:
            self.choose_test_pop()
        self.create_log_daily.function_close_log("test_save_fail")

    def choose_test_pop(self):
        '''
        选择测试次数弹窗,根据table_case_id字典中的数据得到当前选择的测试用例已经测试的次数,进行弹窗显示
        '''
        self.create_log_daily.function_start_log("choose_test_pop")
        try:
            self.table_case_id[self.tableWidget.selectedItems()[
                0].text()]
        except:
            self.create_pop('当前选择测试用例ID为空')
            return
        self.create_log_daily.function_info_log("choose_test_pop", "current selected test_id is {}".format(
            self.tableWidget.selectedItems()[0].text()))
        if self.table_case_id[self.tableWidget.selectedItems()[
                0].text()] == 3:
            self.create_pop("{}测试用例已经进行过三次测试".format(self.tableWidget.selectedItems()[
                0].text()))
        else:
            self.create_pop('请选择测试次数')
        self.create_log_daily.function_close_log("choose_test_pop")

    def change_info(self, err_info):
        '''
        生成测试问题,根据测试用例ID,右侧问题时间,问题输入框的内容生成一条问题记录,修改pushButton_savecount为4
        '''
        self.create_log_daily.function_start_log("change_info")
        if self.label_13.text():
            err_time = self.label_13.text().split(
                ' ')[0][5:] + self.label_13.text().split(' ')[1]
        else:
            err_time = ''
        info = "{}{}\n".format(err_time, err_info)
        self.save_test_info(info)
        self.label_3.setText('')
        self.label_13.setText('')
        self.pushButton_savecount = 4
        self.create_log_daily.function_close_log("change_info")

    def save_test_info(self, info):
        '''
        将生成的测试问题记录在左侧问题显示框中,将右侧问题输入框中的内容清空,设置时间为开始状态
        '''
        self.create_log_daily.function_start_log("save_test_info")
        if not info:
            self.create_pop('请输入问题描述')
            return
        if self.textBrowser_4.toPlainText():
            info = self.textBrowser_4.toPlainText() + info
        else:
            info = info
        self.textBrowser_4.setPlainText(info)
        self.plainTextEdit_2.setPlainText('')
        self.create_log_daily.function_close_log("save_test_info")
        # self.time_count = 1
        # self.timer.start()
        # self.pushButton_8.setText("暂停")
        # self.pushButton_14.setText("暂停")

    def first_test(self):
        self.create_log_daily.function_start_log("first_test")
        # 第一次测试
        try:
            self.tableWidget.selectedItems()[0].text()
        except:
            return
        self.count_for_test(1)
        self.create_log_daily.function_close_log("first_test")

    def second_test(self):
        # 第二测测试
        self.create_log_daily.function_start_log("second_test")
        try:
            self.tableWidget.selectedItems()[0].text()
        except:
            return
        self.count_for_test(2)
        self.create_log_daily.function_close_log("second_test")

    def third_test(self):
        # 第三次测试
        self.create_log_daily.function_start_log("third_test")
        try:
            self.tableWidget.selectedItems()[0].text()
        except:
            return
        self.count_for_test(3)
        self.create_log_daily.function_close_log("third_test")

    def count_for_test(self, count):
        '''
        根据当前选择的table_case_id中的数据判断当前测试用例的测试次数,生成一条信息显示在左侧问题显示框中
        设置pushButton_count为当前测试次数
        '''
        self.create_log_daily.function_start_log("count_for_test")
        try:
            self.table_case_id[self.tableWidget.selectedItems()[0].text()]
        except:
            self.create_pop('当前选择测试用例ID为空')
            return
        if not self.tableWidget.selectedItems()[0].text():
            self.create_log_daily.function_info_log(
                "conut_for_test", "current selected test_id is None")
            self.create_pop('当前选择测试用例ID为空')
            return
        self.create_log_daily.function_info_log(
            "count_for_test", "current selected test_id is {}".format(self.table_case_id[self.tableWidget.selectedItems()[0].text()]))
        self.create_log_daily.function_info_log(
            "count_for_test", "current pushButton_savecount is {}".format(self.pushButton_savecount))
        if self.pushButton_savecount == 0:
            self.create_log_daily.function_info_log(
                "count_for_test", "current selected test_id count is {}".format(self.table_case_id[self.tableWidget.selectedItems()[0].text()]))
            if self.table_case_id[self.tableWidget.selectedItems()[0].text()] == 3:
                self.create_pop("已经完成三次测试")
                return
            if self.table_case_id[self.tableWidget.selectedItems()[0].text()] == count-1:
                self.pushButton_count = count
                self.textBrowser_4.setText("{}第{}次\n".format(
                    self.tableWidget.selectedItems()[0].text(), count))
                self.pushButton_savecount = count
            else:
                self.create_pop('请进行第{}次测试'.format(
                    self.table_case_id[self.tableWidget.selectedItems()[0].text()]+1))
        self.create_log_daily.function_close_log("count_for_test")

    def save_to_excel(self):
        '''
        将测试问题保存在excel文件中,将左侧问题显示框中的内容保存在文件中,根据pushButton_count判断table_case_id中
        的数据,将pushButton_savecount设置为0(初始状态)
        '''
        self.create_log_daily.function_start_log("save_to_excel")
        try:
            self.table_case_id
        except:
            return
        else:
            if not self.table_case_id:
                return
        self.create_log_daily.function_info_log(
            "save_to_excel", "current pushButton_savecount is {}".format(self.pushButton_savecount))
        if self.pushButton_savecount == 4:
            workbook = load_workbook(self.fileName_choose)
            workbook.active
            # 识别excel单元格格式
            workbook.guess_types = True
            new_worksheet = workbook[self.comboBox.currentText()]
            try:
                test_id = re.findall("(.*?)第",
                                     self.textBrowser_4.toPlainText())[0]
            except:
                return
            res_rows_index, res_cols_index = self.get_rows_cols(test_id)
            table_data = xlrd.open_workbook(self.fileName_choose)
            table = table_data.sheet_by_name(self.comboBox.currentText())
            data = table.cell_value(res_rows_index, res_cols_index)
            if res_rows_index and res_cols_index:
                self.create_log_daily.function_info_log(
                    "save_to_excel", "current res_rows_index is {},res_cols_index is {}".format(res_rows_index, res_cols_index))
                info = data + '{}{}\n'.format(
                    self.textBrowser_4.toPlainText(), "{}第{}次".format(test_id, self.table_case_id[test_id]+1))
                new_worksheet.cell(
                    res_rows_index+1, res_cols_index+1).value = info
                workbook.save(self.fileName_choose)
                if self.pushButton_count == 1:
                    self.table_case_id[test_id] = 1
                elif self.pushButton_count == 2:
                    self.table_case_id[test_id] = 2
                elif self.pushButton_count == 3:
                    self.table_case_id[test_id] = 3
                self.textBrowser_4.setText('')
                self.pushButton_savecount = 0
                self.err_list()
            else:
                self.create_pop("添加测试结果失败")
        else:
            self.create_pop('请添加问题')
            return
        self.create_log_daily.function_close_log("save_to_excel")

    def get_rows_cols(self, test_id):
        self.create_log_daily.function_start_log("get_rows_cols")
        # 获取测试结果的列号和测试用例ID列号
        for cols_index in range(self.table.ncols):
            cols_name = self.table.cell_value(0, cols_index)
            if cols_name == "测试用例ID":
                id_cols_index = cols_index
            if cols_name == "测试结果":
                res_cols_index = cols_index
        # 获取测试用例ID行号
        for rows_index in range(self.table.nrows):
            rows_name = self.table.cell_value(rows_index, id_cols_index)
            if rows_name == test_id:
                res_rows_index = rows_index
        self.create_log_daily.function_info_log(
            "get_rows_cols", "current get rows is {},cols is {}".format(res_rows_index, res_cols_index))
        return res_rows_index, res_cols_index
        self.create_log_daily.function_close_log("get_rows_cols")

    def color_change(self, color):
        '''
        改变测试用例的背景颜色
        '''
        self.create_log_daily.function_start_log("color_change")
        self.tableWidget.item(
            self.tableWidget.currentRow(), 2).setForeground(color)
        self.create_log_daily.function_info_log(
            "color_change", "current selected item rows is {}".format(self.tableWidget.currentRow()))
        self.create_log_daily.function_close_log("color_change")
# 第二界面

# 第三界面
    def pull_file(self):
        self.create_log_daily.function_start_log("pull_file")
        self.pull = Pull_File_Views()
        self.pull.show()
        self.create_log_daily.function_close_log("pull_file")

    def choose_push(self):
        self.create_log_daily.function_start_log("choose_push")
        # self.choosepush = Choose_Push_Views()
        # self.choosepush.show()
        self.create_pop('功能暂未开放')
        self.create_log_daily.function_close_log("choose_push")

    def choose_pull(self):
        self.create_log_daily.function_start_log("choose_pull")
        # self.choosepull = Choose_Pull_Views()
        # self.choosepull.show()
        self.create_pop('功能暂未开放')
        self.create_log_daily.function_close_log("choose_pull")

    def open_rviz(self):
        self.create_log_daily.function_start_log("open_rviz")
        self.recordbag = Open_Rviz_Views()
        filepath = Find_File().find_dir_path('zros_dbg_dev_record', '/home/user/')
        self.create_log_daily.function_info_log(
            "open_rviz", "zros_dbg_dev_record path is{}".format(filepath))
        if filepath:
            self.recordbag.show()
        else:
            self.create_pop('没有找到指定的rviz文件')
        self.create_log_daily.function_close_log("open_rvize")

    def brush_soc(self):
        self.create_log_daily.function_start_log("brush_soc")
        # self.brushsoc = Brush_Soc_Views()
        # self.brushsoc.show()
        self.create_pop('功能暂未开放')
        self.create_log_daily.function_close_log("brush_soc")
# 第三界面

# 退出提示
    def closeEvent(self, event):
        # 监听窗口关闭事件
        self.create_log_daily.function_start_log("closeEvent")
        if self.listWidget.count():
            self.problem_record(event)
            problem_res = False
        else:
            problem_res = True
        if self.pushButton_savecount == 4:
            self.case_record(event)
            case_res = False
        else:
            case_res = True
        if problem_res and case_res:
            self.create_log_daily.function_info_log(
                "closeEvent", "This program is about to exit")
            self.create_log_daily.function_close_log("closeEvent")
            os._exit(0)

    def problem_record(self, event):
        self.create_log_daily.function_start_log("problem_record")
        tips_msg = '你还没有保存问题,继续退出吗?'
        self.quit_box(event, tips_msg, self.save_info)
        self.create_log_daily.function_close_log("problem_record")

    def case_record(self, event):
        self.create_log_daily.function_start_log("case_record")
        tips_msg = '你还没有保存用例记录,继续退出吗?'
        self.quit_box(event, tips_msg, self.save_to_excel)
        self.create_log_daily.function_close_log("case_record")

    def quit_box(self, event, tips_msg, handle_fun):
        # 有问题记录没有保存的时候生成退出提醒框
        self.create_log_daily.function_start_log("quit_box")
        self.create_log_daily.function_info_log(
            "quit_box", "current tips_msg is {}".format(tips_msg))
        self.create_log_daily.function_info_log(
            "quit_box", "current handle_function is {}".format(handle_fun))
        quitMsgBox = QMessageBox()
        quitMsgBox.setWindowTitle('确认提示')
        quitMsgBox.setText(tips_msg)
        quitMsgBox.setStandardButtons(
            QMessageBox.Yes | QMessageBox.No | QMessageBox.Ok)
        buttonY = quitMsgBox.button(QMessageBox.Yes)
        buttonY.setText('保存并退出')
        buttonN = quitMsgBox.button(QMessageBox.No)
        buttonN.setText('取消')
        buttonO = quitMsgBox.button(QMessageBox.Ok)
        buttonO.setText('直接退出')
        quitMsgBox.exec_()
        self.create_log_daily.function_info_log(
            "quit_box", "current press keyboard is {}".format(quitMsgBox.clickedButton()))
        if quitMsgBox.clickedButton() == buttonY:
            handle_fun()
            event.accept()
            os._exit(0)
        elif quitMsgBox.clickedButton() == buttonO:
            event.accept()
            os._exit(0)
        else:
            event.ignore()
        self.create_log_daily.function_close_log("quit_box")
# 退出提示


class Pull_File_Views(QDialog, Ui_PullFile):
    '''
        拉取最新的DailyBuild
    '''

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.create_log_daily = CREATE_LOG_DAILY()
        self.initUI()

    def initUI(self):
        self.pushButton.clicked.connect(self.clicked_pull)
        self.timer = QTimer()
        self.timer.timeout.connect(self.pull)
        self.pushButton_2.clicked.connect(self.close)
        self.pushButton_3.clicked.connect(self.choosedir)

    def clicked_pull(self):
        '''
        验证用户选择的文件夹是否是/home/user/,验证成功之后调用pull()函数
        '''
        self.create_log_daily.function_start_log("clicked_pull")
        self.create_log_daily.function_info_log(
            "clicked_pull", "current select filepath is {}".format(self.lineEdit.text()))
        if self.lineEdit.text() == '/home/user':
            pass
        else:
            if self.lineEdit.text()[0:11] == '/home/user/':
                pass
            else:
                self.label_2.setText("请选择/home/user/下的路径")
                self.label_3.setText("")
                self.label_4.setText("")
                return
        self.filepath = self.lineEdit.text()+"/"
        self.label_2.setText("正在拉取文件,请稍候")
        self.label_3.setText("")
        self.label_4.setText("")
        self.timer.start(1)
        self.create_log_daily.function_close_log("clicked_pull")

    def pull(self):
        '''
        运行pull_file.sh,拉取文件,根据pull_file.sh脚本返回的结果判断文件拉取结果
        '''
        self.create_log_daily.function_start_log("pull")
        self.timer.stop()
        try:
            data = os.popen(Generate_File_Path().base_path(
                'Sh/pull_file.sh')+' '+self.filepath)
            data = str(data.read())
        except:
            self.label_3.setText("Unknown Err")
            return
        if data.split('\n')[0] == "连接服务器失败...":
            self.label_3.setText(data.split('\n')[0])
        elif data.split('\n')[0][-2:] == "确定":
            self.label_3.setText(data.split('\n')[1])
            self.label_4.setText("文件拉取完成")
        elif data.split('\n')[0] == "已经有最新的DailyBuild...":
            self.label_3.setText(data.split('\n')[0])
        self.create_log_daily.function_info_log(
            "pull", "current state is {}".format(data.split('\n')[0]))
        self.pushButton_2.setText('完成')
        self.create_log_daily.function_close_log("pull")

    def choosedir(self):
        # 选择文件夹
        self.create_log_daily.function_start_log("choosedir")
        dir_choose = QFileDialog.getExistingDirectory(self,
                                                      "选取文件夹",
                                                      "/home/user/")
        if dir_choose == "":
            return
        self.create_log_daily.function_info_log(
            "choosedir", "current selected dir is {}".format(dir_choose))
        self.lineEdit.setText(dir_choose)
        self.create_log_daily.function_close_log("choosedir")


class Choose_Pull_Views(QDialog, Ui_ChoosePull):
    '''
        选择拉取方式
    '''

    def __init__(self):
        super().__init__()
        self.create_log_daily = CREATE_LOG_DAILY()
        self.setupUi(self)
        self.initUI()

    def initUI(self):
        self.radioButton.toggled.connect(self.choose_method)
        self.radioButton_2.toggled.connect(self.choose_method)
        self.pushButton.clicked.connect(self.check)
        self.pushButton_2.clicked.connect(self.close)

    def choose_method(self):
        '''
        选择拉取方式
        '''
        self.create_log_daily.function_start_log("choose_method")
        self.label_5.setText('')
        if self.radioButton.isChecked():
            self.label_3.setText('当前选择: 通过adb方式拉取文件')
            self.create_log_daily.function_info_log(
                "choose_method", "current selected method is adb")
        else:
            self.label_3.setText('当前选择: 通过scp方式拉取文件')
            self.create_log_daily.function_info_log(
                "choose_method", "current selected method is scp")
        self.create_log_daily.function_close_log("choose_method")

    def check(self):
        '''
        执行所选择拉取方式对应的函数
        '''
        self.create_log_daily.function_start_log("check")
        if self.label_3.text():
            if self.radioButton.isChecked():
                self.adb_pull()
            elif self.radioButton_2.isChecked():
                self.scp_pull()
        else:
            self.create_log_daily.function_info_log(
                "check", "current selected method is None")
            self.label_3.setText('请选择拉取文件的方式')
        self.create_log_daily.function_close_log("check")

    def scp_pull(self):
        self.create_log_daily.function_start_log("scp_pull")
        self.scppull = Scp_Pull_Views()
        self.scppull.show()
        self.close()
        self.create_log_daily.function_close_log("scp_pull")

    def adb_pull(self):
        self.create_log_daily.function_start_log("adb_pull")
        self.adbpull = Adb_Pull_Views()
        self.adbpull.show()
        self.close()
        self.create_log_daily.function_close_log("adb_pull")


class Adb_Pull_Views(QDialog, Ui_AdbPull):
    '''
        Adb拉取文件
    '''

    def __init__(self):
        super().__init__()
        self.create_log_daily = CREATE_LOG_DAILY()
        self.setupUi(self)
        self.initUi()

    def initUi(self):
        self.pushButton.clicked.connect(self.check_input)
        self.pushButton_2.clicked.connect(self.close)
        self.pushButton_3.clicked.connect(self.choosedir)
        self.timer = QTimer()
        self.timer.timeout.connect(self.pull_file)

    def choosedir(self):
        '''
        选择文件夹
        '''
        self.create_log_daily.function_start_log("choosedir")
        self.label_3.setText('')
        self.label_2.setText('')
        dir_choose = QFileDialog.getExistingDirectory(self,
                                                      "选取文件夹",
                                                      "/home/user/")
        if dir_choose == "":
            return
        self.create_log_daily.function_info_log(
            "choosedir", "current selected dir is {}".format(dir_choose))
        self.lineEdit.setText(dir_choose)
        self.create_log_daily.function_close_log("choosedir")

    def check_input(self):
        '''
        检测用户输入信息是否符合要求
        '''
        self.create_log_daily.function_start_log("check_input")
        self.label_3.setText('')
        self.label_2.setText('')
        filepath = self.lineEdit_2.text()
        savepath = self.lineEdit.text()
        if filepath:
            if savepath:
                if filepath[0:6] == '/data/':
                    if savepath[0:11] == '/home/user/':
                        self.label_2.setText("正在拉取文件,请稍等")
                        self.timer.start(1)
                        self.filepath = filepath
                        self.savepath = savepath
                    else:
                        self.label_2.setText("请选择/home/user/下的路径")
                else:
                    self.label_2.setText("请输入/data/下的路径")
            else:
                self.label_2.setText('请选择保存文件路径')
        else:
            self.label_2.setText('请输入拉取文件路径')
        self.create_log_daily.function_close_log("check_input")

    def pull_file(self):
        '''
        拉取文件
        '''
        self.create_log_daily.function_start_log("pull_file")
        self.timer.stop()
        res = subprocess.call('adb pull {} {}'.format(
            self.filepath, self.savepath), shell=True)
        if res:
            self.label_3.setText('拉取失败,检查adb连接或adb权限!!!')
            self.create_log_daily.function_info_log(
                "pull_file", "current pull state is Faild")
        else:
            self.label_3.setText('拉取文件完成')
            self.create_log_daily.function_info_log(
                "pull_file", "current pull state is succeed")

        self.create_log_daily.function_close_log("pull_file")


class Scp_Pull_Views(QDialog, Ui_ScpPull):
    '''
        Scp拉取文件
    '''

    def __init__(self):
        super().__init__()
        self.create_log_daily = CREATE_LOG_DAILY()
        self.setupUi(self)
        self.initUI()

    def initUI(self):
        self.pushButton.clicked.connect(self.pulllog)
        self.pushButton_2.clicked.connect(self.close)
        self.pushButton_3.clicked.connect(self.choosedir)
        self.timer = QTimer()
        self.timer.timeout.connect(self.pull)

    def choosedir(self):
        self.create_log_daily.function_start_log("choosedir")
        self.label_3.setText("")
        self.label_4.setText("")
        dir_choose = QFileDialog.getExistingDirectory(self,
                                                      "选取文件夹",
                                                      "/home/user/")
        if dir_choose == "":
            return
        self.create_log_daily.function_info_log(
            "choosedir", "current selected dir is {}".format(dir_choose))
        self.lineEdit_3.setText(dir_choose)
        self.create_log_daily.function_close_log("choosedir")

    def pulllog(self):
        '''
        检测用户输入信息是否符合要求
        '''
        self.create_log_daily.function_start_log("pulllog")
        self.label_3.setText("")
        self.label_4.setText("")
        host = self.lineEdit.text()
        host_re = "^\d{3}\.\d{3}\.\d{1,}\.\d{3}$"
        hostres = re.findall(host_re, host)
        filepath = self.lineEdit_2.text()
        savepath = self.lineEdit_3.text()
        if hostres:
            if filepath:
                if filepath[0:6] == "/data/":
                    if savepath:
                        if savepath == "/home/user":
                            pass
                        else:
                            if savepath[0:11] == "/home/user/":
                                pass
                            else:
                                self.label_3.setText("请选择/home/user/下的路径")
                                return
                        self.save_logfile(hostres[0], filepath, savepath)
                    else:
                        self.label_3.setText("请选择文件保存路径")
                else:
                    self.label_3.setText("请输入/data/下的路径")
            else:
                self.label_3.setText("请输入文件所在目录")
        else:
            self.label_3.setText('请输入正确的IP地址')
        self.create_log_daily.function_close_log("pulllog")

    def save_logfile(self, host, filepath, savepath):
        '''
        创建timer定时器,拉取文件
        '''
        self.create_log_daily.function_start_log("save_logfile")
        res = subprocess.call("cd {}".format(savepath), shell=True)
        if res:
            self.create_log_daily.function_info_log(
                "save_logfile", "current selected path not is dir")
            self.label_3.setText("保存路径不是一个文件夹")
        else:
            self.create_log_daily.function_info_log(
                "save_logfile", "pulling file, please wait")
            self.label_3.setText('正在拉取文件')
            self.host = host
            self.filepath = filepath
            self.savepath = savepath
            self.timer.start(0.1)
        self.create_log_daily.function_close_log("save_logfile")

    def pull(self):
        self.create_log_daily.function_start_log("pull")
        self.timer.stop()
        if os.path.isfile(self.filepath):
            res = subprocess.call(
                'scp root@{}:{} {}'.format(self.host, self.filepath, self.savepath), shell=True)
        else:
            res = subprocess.call(
                'scp -r root@{}:{} {}'.format(self.host, self.filepath, self.savepath), shell=True)
        if res:
            self.create_log_daily.function_info_log(
                "pull", "current pull state is Faild")
            self.label_4.setText("文件拉取失败,请检查路径是否正确")
        else:
            self.create_log_daily.function_info_log(
                "pull", "current pull state is succeed")
            self.label_4.setText('文件拉取成功')
        self.create_log_daily.function_close_log("pull")


class Choose_Push_Views(QDialog, Ui_ChoosePush):
    '''
        选择推送文件方式
    '''

    def __init__(self):
        super().__init__()
        self.create_log_daily = CREATE_LOG_DAILY()
        self.setupUi(self)
        self.initUI()

    def initUI(self):
        self.radioButton.toggled.connect(self.choose_method)
        self.radioButton_2.toggled.connect(self.choose_method)
        self.pushButton.clicked.connect(self.check)
        self.pushButton_2.clicked.connect(self.close)

    def choose_method(self):
        '''
        选择方式
        '''
        self.create_log_daily.function_start_log("choose_method")
        self.label_5.setText('')
        if self.radioButton.isChecked():
            self.create_log_daily.function_info_log(
                "choose_method", "current selected method is adb")
            self.label_3.setText('当前选择: 通过adb方式推送文件')
        else:
            self.create_log_daily.function_info_log(
                "choose_method", "current selected method is scp")
            self.label_3.setText('当前选择: 通过scp方式推送文件')
        self.create_log_daily.function_close_log("choose_method")

    def check(self):
        '''
        执行所选择方式对应的函数
        '''
        self.create_log_daily.function_start_log("check")
        if self.label_3.text():
            if self.radioButton.isChecked():
                self.adb_push()
            elif self.radioButton_2.isChecked():
                self.scp_push()
        else:
            self.create_log_daily.function_info_log(
                "check", "current select method is None")
            self.label_3.setText('请选择推送文件的方式')
        self.create_log_daily.function_close_log("check")

    def adb_push(self):
        self.create_log_daily.function_start_log("adb_push")
        self.adbpush = Adb_Push_Views()
        self.adbpush.show()
        self.close()
        self.create_log_daily.function_close_log("adb_push")

    def scp_push(self):
        self.create_log_daily.function_start_log("scp_push")
        self.scppush = Scp_Push_Views()
        self.scppush.show()
        self.close()
        self.create_log_daily.function_close_log("scp_push")


class Adb_Push_Views(QDialog, Ui_AdbPush):
    '''
        Adb推送文件
    '''

    def __init__(self):
        super().__init__()
        self.create_log_daily = CREATE_LOG_DAILY()
        self.setupUi(self)
        self.initUI()

    def initUI(self):
        self.pushButton_3.clicked.connect(self.choosefile)
        self.pushButton_2.clicked.connect(self.close)
        self.pushButton.clicked.connect(self.pushfile)
        self.timer = QTimer()
        self.timer.timeout.connect(self.push)

    def choosefile(self):
        self.create_log_daily.function_start_log("choosefile")
        self.label_3.setText('')
        self.label_2.setText('')
        files, ok1 = QFileDialog.getOpenFileName(self,
                                                 "文件选择",
                                                 "/home/user/",
                                                 "All Files (*);;Text Files (*.txt)")
        self.create_log_daily.function_info_log(
            "choosefile", "current selected file is {}".format(files))
        self.lineEdit.setText(files)
        self.create_log_daily.function_close_log("choosefile")

    def pushfile(self):
        '''
        检测用户输入信息是否符合要求
        '''
        self.create_log_daily.function_start_log("pushfile")
        pushpath = self.lineEdit_2.text()
        filepath = self.lineEdit.text()
        self.label_2.setText('')
        self.label_3.setText('')
        if pushpath:
            if filepath:
                if pushpath[0:6] == '/data/':
                    res = subprocess.call(
                        '[ -d {} ]'.format(filepath), shell=True)
                    if res:
                        self.label_2.setText("正在推送文件,请稍等")
                        self.pushpath = pushpath
                        self.filepath = filepath
                        self.timer.start(1)
                    else:
                        self.label_2.setText('请选择正确的文件')
                else:
                    self.label_2.setText("请输入/data/下的路径")
            else:
                self.label_2.setText("请选择本地文件")
        else:
            self.label_2.setText("请输入目标路径")
        self.create_log_daily.function_close_log("pushfile")

    def push(self):
        '''
        推送文件
        '''
        self.create_log_daily.function_start_log("push")
        self.timer.stop()
        res = subprocess.call('adb push {} {}'.format(
            self.filepath, self.pushpath), shell=True)
        if res:
            self.create_log_daily.function_info_log(
                "push", "current push state is Faild")
            self.label_3.setText('文件推送失败,请检查adb连接或adb权限!!!')
        else:
            self.create_log_daily.function_info_log(
                "push", "current push state is succeed")
            self.label_3.setText('文件推送成功')
        self.create_log_daily.function_close_log("push")


class Scp_Push_Views(QDialog, Ui_ScpPush):
    '''
        Scp推送文件
    '''

    def __init__(self):
        super().__init__()
        self.create_log_daily = CREATE_LOG_DAILY()
        self.setupUi(self)
        self.initUI()

    def initUI(self):
        self.pushButton.clicked.connect(self.pushfile)
        self.pushButton_3.clicked.connect(self.choosefile)
        self.pushButton_2.clicked.connect(self.close)
        self.timer = QTimer()
        self.timer.timeout.connect(self.push)

    def pushfile(self):
        '''
        检测用户输入信息是否符合要求
        '''
        self.create_log_daily.function_start_log("pushfile")
        self.label_3.setText("")
        self.label_4.setText("")
        host = self.lineEdit.text()
        pushpath = self.lineEdit_2.text()
        filepath = self.lineEdit_3.text()
        host_re = "^\d{3}\.\d{3}\.\d{1,}\.\d{3}$"
        hostres = re.findall(host_re, host)
        if hostres:
            if pushpath[0:6] == "/data/":
                if filepath[0:11] == "/home/user/":
                    self.label_3.setText("正在推送文件,请稍等")
                    self.host = hostres[0]
                    self.pushpath = pushpath
                    self.filepath = filepath
                    self.timer.start(1)
                else:
                    self.label_3.setText("请选择/home/user/下的文件")
            else:
                self.label_3.setText("请输入/data/路径下的目标路径")
        else:
            self.label_3.setText("请输入正确的IP")
        self.create_log_daily.function_close_log("pushfile")

    def choosefile(self):
        self.create_log_daily.function_start_log("choosefile")
        self.label_3.setText("")
        self.label_4.setText("")
        files, ok1 = QFileDialog.getOpenFileName(self,
                                                 "文件选择",
                                                 "/home/user/",
                                                 "All Files (*);;Text Files (*.txt)")
        self.create_log_daily.function_info_log(
            "choosefile", "current selected file is {}".format(files))
        self.lineEdit_3.setText(files)
        self.create_log_daily.function_close_log("choosefile")

    def push(self):
        '''
        推送文件
        '''
        self.create_log_daily.function_start_log("push")
        self.timer.stop()
        if os.path.isfile(self.filepath):
            res = subprocess.call(
                "scp {} root@{}:{}".format(self.filepath, self.host, self.pushpath), shell=True)
        else:
            res = subprocess.call(
                "scp -r {} root@{}:{}".format(self.filepath, self.host, self.pushpath), shell=True)
        if res:
            self.create_log_daily.function_info_log(
                "push", "current push state is Faild")
            self.label_4.setText("文件推送失败,请检查目标路径或本地文件是否正确")
        else:
            self.create_log_daily.function_info_log(
                "push", "current push state is succeed")
            self.label_4.setText("文件推送成功")
        self.create_log_daily.function_close_log("push")


class Open_Rviz_Views(QDialog, Ui_OpenRviz):
    '''
        打开Rviz
    '''

    def __init__(self):
        super().__init__()
        self.create_log_daily = CREATE_LOG_DAILY()
        self.setupUi(self)
        self.initUi()

    def initUi(self):
        self.ip = ''
        self.filepath = Find_File().find_dir_path('zros_dbg_dev_record', '/home/user/')
        self.pushButton.clicked.connect(self.record_bag)
        self.pushButton_2.clicked.connect(self.close)

    def check_ip(self):
        '''
        检测输入的ip地址是否符合要求
        '''
        self.create_log_daily.function_start_log("check_ip")
        self.label_2.setText('')
        address = self.lineEdit.text()
        if not address:
            self.label_2.setText('请输入ip地址')
            return 0
        address_re = "^\d{3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$"
        res = re.findall(address_re, address)
        if not res:
            self.create_log_daily.function_info_log(
                "check_ip", "current input ip format is Faild")
            self.label_2.setText('输入ip格式错误')
            return 0
        self.ip = res[0]
        self.create_log_daily.function_info_log(
            "check_ip", "current input ip is {}".format(self.ip))
        self.create_log_daily.function_close_log("check_ip")
        return 1

    def record_bag(self):
        '''
            修改zros_dbg_dev_record中localization.launch中的ip
            执行start_all.sh 传入参数
        '''
        self.create_log_daily.function_start_log("record_bag")
        if not self.check_ip():
            return
        launchfile = Find_File().find_file_path(
            'localization.launch', self.filepath)[0]
        with open(launchfile, 'r') as f:
            server_address = re.findall(
                '"server_address">(.*)</rosparam>', f.read())
        self.create_log_daily.function_info_log(
            "record_bag", "current input ip is {}".format(self.ip))
        subprocess.call(
            "sed -i 's/{}/{}/g' {}".format(server_address[0], self.ip, launchfile), shell=True)
        os.popen(Generate_File_Path().base_path('Sh/start_all.sh')+' ' +
                 Generate_File_Path().base_path('Sh/record_bag.sh')+' '+Generate_File_Path().base_path('Sh/rviz_e.sh')+' '+self.filepath+' '+self.ip)
        sleep(1)
        self.close()
        self.create_log_daily.function_close_log("record_bag")


class Brush_Soc_Views(QDialog, Ui_BrushSoc):
    '''
        刷写Soc版本
    '''

    def __init__(self):
        super().__init__()
        self.create_log_daily = CREATE_LOG_DAILY()
        self.setupUi(self)
        self.tarfilepath = ''
        self.address = ''
        self.default_existent = 1
        self.timer = QTimer()
        self.initUi()

    def initUi(self):
        self.pushButton.clicked.connect(self.choose_file)
        self.pushButton_2.clicked.connect(self.check_pushbutton_text)
        self.pushButton_3.clicked.connect(self.close)
        self.timer.timeout.connect(self.brush_soc)

    def choose_file(self):
        self.create_log_daily.function_start_log("choose_file")
        files, ok1 = QFileDialog.getOpenFileName(self,
                                                 "文件选择",
                                                 "/home/user/",
                                                 "All Files (*);;Text Files (*)")
        self.create_log_daily.function_info_log(
            "choose_file", "current selected file is {}".format(files))
        self.tarfilepath = files
        self.create_log_daily.function_close_log("choose_file")

    def check_pushbutton_text(self):
        self.create_log_daily.function_start_log("check_pushbutton_text")
        self.create_log_daily.function_info_log(
            "check_pushbutton_text", "current pushbutton_text is {}".format(self.pushButton_2.text()))
        if self.pushButton_2.text() == "刷写soc":
            self.check_ip_tar()
        elif self.pushButton_2.text() == "立即重启":
            self.reboot_now()
        else:
            pass
        self.create_log_daily.function_close_log("check_pushbutton_text")

    def check_ip_tar(self):
        '''
        检测用户输入的ip地址和选择的soc版本文件
        创建timer计时器
        '''
        self.create_log_daily.function_start_log("check_ip_tar")
        self.label_4.setText('')
        self.address = self.lineEdit.text()
        if not self.address:
            self.label_4.setText('ip不能为空')
            return
        address_re = "^\d{3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$"
        res = re.findall(address_re, self.address)
        if not res:
            self.create_log_daily.function_info_log(
                "check_ip_tar", "current input ip format is Faild")
            self.label_4.setText('输入ip格式错误')
            return
        if self.tarfilepath[-7:] != '.tar.gz':
            self.create_log_daily.function_info_log(
                "check_ip_tar", "current selected zros_file extension not is .tar.gz")
            self.label_4.setText('请选择.tar.gz文件')
            return
        self.label_4.setText('正在刷写dailybuild,请稍等')
        self.timer.start(1)
        self.create_log_daily.function_close_log("check_ip_tar")

    def brush_soc(self):
        '''
        检测板子中是否有原始文件
        '''
        self.timer.stop()
        self.create_log_daily.function_start_log("brush_soc")
        address = 'root@{}'.format(self.address)
        self.create_log_daily.function_info_log(
            "current address is {}".format(address))
        Mkdir_Path_Views().mkdir_dir_path('/home/user/Data/car_instance/')
        subprocess.call('cd /home/user/Data/car_instance/', shell=True)
        res = subprocess.call(
            'timeout 3 ssh {} "rmdir /data/zros/"'.format(address), shell=True)
        if res:
            # 板子中/data/zros/文件夹不为空
            self.default_existent = 1
            res = subprocess.call(
                'timeout 2 adb push {} /usr/bin/'.format(Generate_File_Path().base_path('Sh/killallnodes')), shell=True)
            if not res:
                subprocess.call('timeout 2 adb push {} /usr/bin/'.format(
                    Generate_File_Path().base_path('Sh/get_node_list')), shell=True)
                res = subprocess.call(
                    'timeout 3 scp -p {}:/data/zros/res/car_instance/default* /home/user/Data/car_instance/'.format(address), shell=True)
                if not res:
                    self.default_existent = 1
                else:
                    self.default_existent = 0
                subprocess.call(
                    'timeout 2 ssh {} "killallnodes;cd /data/zros/;rm -r *"'.format(address), shell=True)
                res = self.push_tar(address)
                if res:
                    print('\nError')
                else:
                    self.label_4.setText("刷写完成")
                    print('\nSuccessful')
                    self.pushButton_2.setText('立即重启')
                    self.pushButton_3.setText('稍后重启')
            else:
                self.label_4.setText('请检查adb连接')
        else:
            # 板子中/data/zros/文件夹为空
            self.default_existent = 0
            res = subprocess.call(
                'timeout 3 ssh {} "mkdir /data/zros/"'.format(address), shell=True)
            res = self.push_tar(address)
            if res:
                print('\nError')
            else:
                self.label_4.setText('刷写完成')
                print('\nSuccessful')
                self.pushButton_2.setText('立即重启')
                self.pushButton_3.setText('稍后重启')
        self.create_log_daily.function_close_log("brush_soc")

    def push_tar(self, address):
        '''
        刷写soc版本
        '''
        self.create_log_daily.function_start_log("push_tar")
        res = subprocess.call(
            'adb push -p {} /data/zros/'.format(self.tarfilepath), shell=True)
        if not res:
            res = subprocess.call(
                'ssh {} "cd /data/zros/;tar -zxvf *"'.format(address), shell=True)
        else:
            self.create_log_daily.function_info_log(
                "push_tar", "push dailybuild Faild, please check adb connect")
            self.label_4.setText('推送dailybuild失败,请检查adb连接')
            return 1
        if not res:
            res = self.push_default(address)
            return res
        else:
            self.create_log_daily.function_info_log(
                "push_tar", "unzip dailybuild Faild, please check system memory")
            self.label_4.setText('解压dailybuild失败,请查看板子内存是否足够')
            return 1
        self.create_log_daily.function_close_log("push_tar")

    def push_default(self, address):
        self.create_log_daily.function_start_log("push_default")
        if self.default_existent:
            res = subprocess.call(
                'timeout 3 scp /home/user/Data/car_instance/default* {}:/data/zros/res/car_instance/'.format(address), shell=True)
            if not res:
                return 0
            else:
                self.create_log_daily.function_info_log(
                    "push_default", "replace default.xml and defaultDevice Faild")
                self.label_4.setText('替换default.xml和defaultDevice文件失败')
                return 1
        else:
            self.create_log_daily.function_info_log(
                "push_default", "please manually replace default.xml and defaultDevice")
            self.label_4.setText('请手动替换default.xml和defaultDevice')
            return 0
        self.create_log_daily.function_close_log("push_default")

    def reboot_now(self):
        self.create_log_daily.function_start_log("reboot_now")
        address = 'root@{}'.format(self.address)
        self.create_log_daily.function_info_log(
            "reboot_now", "current address is {}".format(address))
        subprocess.call(
            'timeout 2 ssh {} "/sbin/reboot"'.format(address), shell=True)
        self.close()
        self.create_log_daily.function_close_log("reboot_now")


class Mkdir_Path_Views(object):
    '''
        创建文件夹
    '''

    def __init__(self):
        self.create_log_daily = CREATE_LOG_DAILY()

    def mkdir_dir_path(self, dir_path):
        '''
        dir_path:需要创建的文件夹路径
        return: 1 创建成功 0 创建失败
        '''
        self.create_log_daily.function_start_log("mkdir_dir_path")
        self.create_log_daily.function_info_log(
            "mkdir_dir_path", "current mkdir dir path is {}".format(dir_path))
        if subprocess.call('mkdir -p {}'.format(dir_path), shell=True):
            return 0
        else:
            return 1
        self.create_log_daily.function_close_log("mkdir_dir_path")

    def touch_file_path(self, file_path):
        '''
        file_path:需要创建的文件夹路径
        return: 1 创建成功 0 创建失败
        '''
        self.create_log_daily.function_start_log("touch_file_path")
        self.create_log_daily.function_info_log(
            "touch_file_path", "current touch file path is {}".format(file_path))
        if subprocess.call('touch {}'.format(file_path), shell=True):
            return 0
        else:
            return 1
        self.create_log_daily.function_close_log("touch_file_path")


class Find_File(object):
    '''
    查找文件
    '''

    def __init__(self):
        self.create_log_daily = CREATE_LOG_DAILY()
        self.file_path_list = []

    def find_file_path(self, filename, startdir):
        '''
        filename:需要查找的文件名;startdir:开始查找的起始路径
        return:查找到的所有文件路径列表
        '''
        self.create_log_daily.function_start_log("find_file_path")
        for root, dirs, names in os.walk(startdir, topdown=False):
            for fname in names:
                if fname == filename:
                    filepath = os.path.join(root, fname)
                    self.create_log_daily.function_info_log(
                        "find_file_path", "current find file path is {}".format(filepath))
                    self.file_path_list.append(filepath)
        self.create_log_daily.function_close_log("find_file_path")
        return self.file_path_list

    def find_dir_path(self, dirname, startdir):
        '''
        dirname:需要查找的文件夹名;startdir:开始查找的起始路径
        retrun:查找到的所有文件夹路径的列表
        '''
        self.create_log_daily.function_start_log("find_dir_path")
        for root, dirs, names in os.walk(startdir):
            for dname in dirs:
                if dname == dirname:
                    dirpath = os.path.join(root, dname)
                    self.create_log_daily.function_info_log(
                        "find_dir_path", "current find dir path is {}".format(dirpath))
                    return dirpath
        self.create_log_daily.function_close_log("find_dir_path")


class Generate_File_Path(object):
    '''
        生成文件当前运行路径
    '''

    def __init__(self):
        self.create_log_daily = CREATE_LOG_DAILY()

    def base_path(self, path):
        '''
        path:文件当前位置
        retrun:文件在不同运行环境中的路径
        '''
        self.create_log_daily.function_start_log("base_path")
        if getattr(sys, 'frozen', None):
            basedir = sys._MEIPASS
        else:
            basedir = os.path.dirname(
                os.path.dirname(os.path.abspath(__file__)))
        self.create_log_daily.function_info_log(
            "base_path", "current working directory is {}".format(basedir))
        self.create_log_daily.function_close_log("base_path")
        return os.path.join(basedir, path)


class CREATE_LOG_DAILY(object):
    '''
        打印函数log日志
    '''

    def function_start_log(self, function_name):
        print("[{}][dug][{}]FUNCTION {} IS START".format(time.strftime("%Y-%m-%d %H:%M:%S",
                                                                       time.gmtime()), function_name, function_name))

    def function_info_log(self, function_name, function_info):
        print("[{}][dug][{}]{}".format(time.strftime("%Y-%m-%d %H:%M:%S",
                                                     time.gmtime()), function_name, function_info))

    def function_close_log(self, function_name):
        print("[{}][dug][{}]{}".format(time.strftime("%Y-%m-%d %H:%M:%S",
                                                     time.gmtime()), function_name, "FUNCTION {} IS CLOSE".format(function_name)))
